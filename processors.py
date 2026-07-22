# -*- coding: utf-8 -*-
"""
processors.py v3.0 - ERP 資料解析 -> 經銷商歸戶 -> Excel 報表產生
v3.0 變更: 折數標籤格式修正、額外欄位完整支援、Sheet name 對應修正
"""

from collections import defaultdict
from datetime import datetime, date
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl, re, io, json, os

from config import (
    CATEGORY_MAP, CATEGORY_ORDER, CATEGORY_COL,
    ERP_COLUMN_DEF, SHEET_HEADERS, INVALID_SHEET_CHARS,
    MAX_SHEET_NAME_LEN, DEFAULT_DISCOUNT_RATES, PAGE_TITLE,
    REGION_GROUP, SHEET_ORDER, DISPLAY_NAME_MAP,
    SHEET_TO_STD, STD_TO_SHEET, format_rate_display,
    BASE_DIR, load_distributor_extras,
)

# Load R3 rate labels (exact strings from target file)
def _load_r3_labels():
     path = os.path.join(BASE_DIR, "r3_rate_labels.json")
     if not os.path.exists(path):
         return {}
     with open(path, "r", encoding="utf-8") as f:
         return json.load(f)


# ============================================================
# 工具函式
# ============================================================

def safe_float(val):
     if val is None:
         return 0.0
     try:
         return float(val)
     except (ValueError, TypeError):
         return 0.0

def parse_date(val):
     if val is None:
         return None
     if isinstance(val, datetime):
         return val.replace(tzinfo=None) if val.tzinfo else val
     if isinstance(val, date):
         return datetime(val.year, val.month, val.day)
     s = str(val).strip()
     for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
         try:
             return datetime.strptime(s, fmt)
         except ValueError:
             continue
     return None

def get_category(product_code):
     if not product_code or len(str(product_code)) < 2:
         return "其他"
     prefix = str(product_code)[:2]
     return CATEGORY_MAP.get(prefix, "其他")

def safe_sheet_name(name):
     safe = str(name)
     for ch in INVALID_SHEET_CHARS:
         safe = safe.replace(ch, "")
     return safe[:MAX_SHEET_NAME_LEN]

def sheet_ref(sname):
     SQ = chr(39)
     needs_quote = any(c in sname for c in SQ + " () ")
     if needs_quote:
         escaped = sname.replace(SQ, SQ + SQ)
         return SQ + escaped + SQ
     return sname

def most_common_date(dates):
     if not dates:
         return None
     c = defaultdict(int)
     for d in dates:
         if d:
             c[d] += 1
     return max(c, key=c.get) if c else None

def _get_rate_for_round(rate_val):
     """產生 ROUND 公式用的折扣率字串"""
     if rate_val is None or rate_val == 0:
         return "0"
     rv = float(rate_val)
     s = f"{rv:.3f}".rstrip("0")
     return s


# ============================================================
# ERP 資料解析
# ============================================================

def _parse_row_from_list(cells, code_to_dist):
     inv_no = str(cells[ERP_COLUMN_DEF["invoice_no"]]).strip() if len(cells) > ERP_COLUMN_DEF["invoice_no"] and cells[ERP_COLUMN_DEF["invoice_no"]] else ""
     cust_code = str(cells[ERP_COLUMN_DEF["cust_code"]]).strip() if len(cells) > ERP_COLUMN_DEF["cust_code"] and cells[ERP_COLUMN_DEF["cust_code"]] else ""
     if not inv_no or not cust_code:
         return None, None, False
     d_val = parse_date(cells[ERP_COLUMN_DEF["date"]]) if len(cells) > ERP_COLUMN_DEF["date"] and cells[ERP_COLUMN_DEF["date"]] else None
     product_code_raw = str(cells[ERP_COLUMN_DEF["product_code"]]).strip() if len(cells) > ERP_COLUMN_DEF["product_code"] and cells[ERP_COLUMN_DEF["product_code"]] else ""
     amount = safe_float(cells[ERP_COLUMN_DEF["amount"]]) if len(cells) > ERP_COLUMN_DEF["amount"] and cells[ERP_COLUMN_DEF["amount"]] is not None else 0.0
     if amount == 0 and ("運費" in product_code_raw or product_code_raw.startswith("002")):
         return None, d_val, False
     distributor = code_to_dist.get(cust_code)
     cust_name_raw = str(cells[ERP_COLUMN_DEF["cust_name"]]).strip() if len(cells) > ERP_COLUMN_DEF["cust_name"] and cells[ERP_COLUMN_DEF["cust_name"]] else cust_code
     is_unmatched = (distributor is None)
     return {"date": d_val, "inv_no": inv_no, "cust_code": cust_code, "cust_name": cust_name_raw, "product_code": product_code_raw, "amount": amount, "distributor": distributor}, d_val, is_unmatched

def _build_invoice_records(rows_data, code_to_dist):
     inv_records = defaultdict(lambda: {"date": None, "cust_name": "", "cust_code": "", "cats": defaultdict(float), "distributor": None})
     all_dates = []
     unmatched = []
     total_read = 0
     for rec, d_val, is_unmatched in rows_data:
         if rec is None:
             continue
         total_read += 1
         if d_val:
             all_dates.append(d_val)
         if is_unmatched and not any(u[0] == rec["inv_no"] for u in unmatched):
             unmatched.append((rec["inv_no"], rec["cust_code"]))
         key = rec["inv_no"]
         r = inv_records[key]
         if r["date"] is None:
             r["date"] = rec["date"]
             r["cust_name"] = rec["cust_name"]
             r["cust_code"] = rec["cust_code"]
             r["distributor"] = rec["distributor"]
         cat = get_category(rec["product_code"])
         r["cats"][cat] += rec["amount"]
     return dict(inv_records), unmatched, total_read, all_dates

def parse_erp_paste(text, code_to_dist):
     lines_data = text.strip().split("\n")
     rows_data = []
     header_keywords = {"日期(轉換)", "進銷單號", "客戶供應商代號", "產品代號", "品名"}
     for line in lines_data:
         cells = line.split("\t")
         if len(cells) < 2:
             continue
         line_vals = {str(c).strip() for c in cells}
         if line_vals & header_keywords:
             continue
         result, d_val, is_unmatched = _parse_row_from_list(cells, code_to_dist)
         rows_data.append((result, d_val, is_unmatched))
     return _build_invoice_records(rows_data, code_to_dist)

def parse_erp_excel(file_bytes, code_to_dist):
     wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
     ws = wb.active
     rows_data = []
     for row in ws.iter_rows(min_row=2):
         cells = [c.value for c in row]
         if not any(v is not None and str(v).strip() for v in cells[:4]):
             continue
         result, d_val, is_unmatched = _parse_row_from_list(cells, code_to_dist)
         rows_data.append((result, d_val, is_unmatched))
     wb.close()
     return _build_invoice_records(rows_data, code_to_dist)


# ============================================================
# 經銷商歸戶
# ============================================================

def aggregate_by_distributor(inv_records):
     dist_groups = defaultdict(dict)
     unmatched_count = 0
     for inv_no, record in inv_records.items():
         distributor = record.get("distributor")
         if distributor is None:
             unmatched_count += 1
             continue
         cats = {k: round(v, 2) for k, v in record["cats"].items()}
         dist_groups[distributor][inv_no] = {"date": record["date"], "cust_name": record["cust_name"], "cust_code": record["cust_code"], "cats": cats}
     return dict(dist_groups), unmatched_count


# ============================================================
# Excel 報表產生 - 彙總頁
# ============================================================

def _init_summary_sheet(ws, dist_info):
     ws.cell(row=2, column=3, value="比率")
     for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
         col = CATEGORY_COL[cat_name]
         cl = get_column_letter(col)
         ws.cell(row=2, column=col, value=f"={cl}3/I3")
     ws.cell(row=2, column=9, value=1)
     for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
         col = CATEGORY_COL[cat_name] + 6
         cl = get_column_letter(col)
         ws.cell(row=2, column=col, value=f"={cl}3/P3")
     ws.cell(row=2, column=15, value="=O3/P3")
     ws.cell(row=2, column=16, value=1)
     ws.cell(row=2, column=18, value="退貨")
     ws.cell(row=2, column=19, value="牌+退貨")
     ws.cell(row=2, column=20, value="退貨%")
     ws.cell(row=2, column=21, value="北區合計")
     ws.cell(row=2, column=22, value="中區合計")
     ws.cell(row=2, column=23, value="南區合計")
     ws.cell(row=2, column=24, value="小客戶合計")

     ws.cell(row=3, column=2, value="=P3/I3")
     ws.cell(row=3, column=3, value="合計")
     for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
         col = CATEGORY_COL[cat_name]
         cl = get_column_letter(col)
         ws.cell(row=3, column=col, value=f"=SUM({cl}6:{cl}40)")
     ws.cell(row=3, column=9, value="=SUM(I6:I40)")
     for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
         col = CATEGORY_COL[cat_name] + 6
         cl = get_column_letter(col)
         ws.cell(row=3, column=col, value=f"=SUM({cl}6:{cl}40)")
     ws.cell(row=3, column=15, value="=SUM(O6:O40)")
     ws.cell(row=3, column=16, value="=SUM(P6:P40)")
     ws.cell(row=3, column=18, value="=SUM(R6:R40)")
     ws.cell(row=3, column=19, value="=U3+V3+W3+X3")
     ws.cell(row=3, column=20, value="=SUM(R3/I3)")

     target_orders = {"北區": [8,10,15,7,16,39,33,34], "中區": [9,11,6,12,17,20,23,29,30],
                       "南區": [13,14,18,19,21,37,38,31,32,35,36], "小客戶": [40,22,24,25,26,27,28]}
     col_map = {"北區": 21, "中區": 22, "南區": 23, "小客戶": 24}
     for rname, rows in target_orders.items():
         sum_args = "+".join("S{}".format(r) for r in rows)
         ws.cell(row=3, column=col_map[rname], value=f"=SUM({sum_args})")

     ws.cell(row=4, column=3, value="名稱")
     ws.cell(row=4, column=4, value="牌   價")
     ws.cell(row=4, column=10, value="實   價")
     ws.cell(row=4, column=18, value="牌價")
     ws.cell(row=4, column=19, value=5500)
     ws.cell(row=4, column=23, value=5500)

     cat_labels = {"單斤": "單味斤裝", "複斤": "複方斤裝", "單濃": "單味濃縮", "複濃": "複方濃縮", "O.T.C.": "O.T.C成藥"}
     ws.cell(row=5, column=2, value="折數")
     for cat_name, col_num in CATEGORY_COL.items():
         ws.cell(row=5, column=col_num, value=cat_labels.get(cat_name, cat_name))
     ws.cell(row=5, column=9, value="合計")
     for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
         col = CATEGORY_COL[cat_name] + 6
         ws.cell(row=5, column=col, value=cat_labels.get(cat_name, cat_name))
     ws.cell(row=5, column=15, value="其     他")
     ws.cell(row=5, column=16, value="合計")
     ws.cell(row=5, column=18, value="合計")
     ws.cell(row=5, column=19, value="合計")
     ws.cell(row=5, column=21, value="月底帳單")
     ws.cell(row=5, column=23, value="=SUM(U1:X1)")
     ws.cell(row=5, column=26, value="=SUM(W5-S3)")

     for i, sname in enumerate(SHEET_ORDER):
         dr = 6 + i
         sr = sheet_ref(sname)
         display_name = DISPLAY_NAME_MAP.get(sname, sname)
         ws.cell(row=dr, column=2, value=f"=P{dr}/I{dr}")
         ws.cell(row=dr, column=3, value=display_name)
         for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
             col = CATEGORY_COL[cat_name]
             cl = get_column_letter(col)
             ws.cell(row=dr, column=col, value=f"={sr}!{cl}2")
         d_cl = get_column_letter(CATEGORY_COL["單斤"])
         h_cl = get_column_letter(CATEGORY_COL["O.T.C."])
         ws.cell(row=dr, column=9, value=f"=SUM({d_cl}{dr}:{h_cl}{dr})")
         for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
             sc = CATEGORY_COL[cat_name]
             dc = sc + 6
             cs = get_column_letter(sc)
             cd = get_column_letter(dc)
             ws.cell(row=dr, column=dc, value=f"={sr}!{cs}4")
         ws.cell(row=dr, column=15, value=f"={sr}!I4")
         j_cl = get_column_letter(CATEGORY_COL["單斤"]+6)
         ws.cell(row=dr, column=16, value=f"=SUM({j_cl}{dr}:O{dr})")
         ws.cell(row=dr, column=18, value=0)
         ws.cell(row=dr, column=19, value=f"=R{dr}+I{dr}")
         ws.cell(row=dr, column=20, value=f"=SUM(R{dr}/I{dr})")

     return ws


# ============================================================
# Excel 報表產生 - 經銷商個別頁
# ============================================================

def _write_distributor_extras(ws, extras):
     for cell_ref, (vtype, vvalue) in extras.items():
         col_letter = "".join(c for c in cell_ref if c.isalpha())
         row_num = int("".join(c for c in cell_ref if c.isdigit()))
         try:
             ws.cell(row=row_num, column=column_index_from_string(col_letter), value=vvalue)
         except Exception:
             pass

def create_new_distributor_sheet(wb, std_name, dist_info, month_display):
     info = dist_info.get(std_name, {})
     sname_lookup = info.get("sheet_name", STD_TO_SHEET.get(std_name, std_name))
      # Match to SHEET_ORDER for exact name
     matched_sname = None
     for so in SHEET_ORDER:
         if safe_sheet_name(so).strip() == safe_sheet_name(sname_lookup).strip():
             matched_sname = so
             break
     sname = matched_sname if matched_sname else safe_sheet_name(sname_lookup)

     company_full = info.get("company_full", std_name)
     rates = info.get("discount_rates", DEFAULT_DISCOUNT_RATES.copy())

     ws = wb.create_sheet(title=sname)

      # R1
     ws.cell(row=1, column=1, value="請款：")
     ws.cell(row=1, column=2, value=month_display)
     ws.cell(row=1, column=3, value="經銷商:")
     ws.cell(row=1, column=4, value=company_full)

      # R2 - SUM formulas (initial range 7:1000)
     ws.cell(row=2, column=2, value="合  計")
     ws.cell(row=2, column=3, value="=D2+E2+F2+G2+H2")
     for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
         col = CATEGORY_COL[cat_name]
         cl = get_column_letter(col)
         ws.cell(row=2, column=col, value=f"=SUM({cl}7:{cl}1000)")
     ws.cell(row=2, column=9, value="=SUM(I7:I1000)")

      # R3 - Use exact rate labels from JSON (matching target file)
     r3_labels = _load_r3_labels()
     label_set = r3_labels.get(sname, {})
     for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
         col = CATEGORY_COL[cat_name]
         ws.cell(row=3, column=col, value=label_set.get(cat_name, format_rate_display(rates.get(cat_name, 0))))

      # R4 - ROUND formulas
     ws.cell(row=4, column=2, value="總   計")
     for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
         col = CATEGORY_COL[cat_name]
         rate_val = rates.get(cat_name, 0)
         cl = get_column_letter(col)
         rs = _get_rate_for_round(rate_val)
         ws.cell(row=4, column=col, value=f"=ROUND({cl}2*{rs},0)")
     ws.cell(row=4, column=9, value="=ROUND(SUM(I2:I2),0)")

      # R5
     ws.cell(row=5, column=1, value="銷貨")
     ws.cell(row=5, column=2, value="應收帳款")
     ws.cell(row=5, column=3, value="=D4+E4+F5+H4+I4")
     ws.cell(row=5, column=6, value="=ROUND(F4+G4,0)")

      # R6 - Headers
     for idx, hdr in enumerate(SHEET_HEADERS):
         ws.cell(row=6, column=idx + 1, value=hdr)

      # Extras (J-W columns from extras JSON)
     all_extras = load_distributor_extras()
     extras = all_extras.get(sname, {})
     if extras:
         _write_distributor_extras(ws, extras)

     return ws


# ============================================================
# 資料追加與公式更新
# ============================================================

def find_last_data_row(ws):
     last_row = 6
     for r in range(7, ws.max_row + 1):
         if ws.cell(row=r, column=2).value is not None:
             last_row = r
     return last_row

def append_data_to_sheet(ws, records):
     current_last = find_last_data_row(ws)
     added = 0
     skipped = 0
     for inv_no, record in records.items():
         existing_inv = False
         for check_row in range(7, current_last + 1):
             if str(ws.cell(row=check_row, column=2).value) == inv_no:
                 existing_inv = True
                 skipped += 1
                 break
         if existing_inv:
             continue
         r = current_last + 1
         added += 1
         cats = record["cats"]
         if record["date"]:
             ws.cell(row=r, column=1, value=record["date"])
         ws.cell(row=r, column=2, value=inv_no)
         ws.cell(row=r, column=3, value=record["cust_name"])
         for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
             col = CATEGORY_COL[cat_name]
             amt = cats.get(cat_name, 0)
             if amt > 0:
                 ws.cell(row=r, column=col, value=int(round(amt)))
         current_last = r
     return current_last, added, skipped

def update_sheet_formulas(ws, last_row):
     for cat_name in ["單斤", "複斤", "單濃", "複濃", "O.T.C."]:
         col = CATEGORY_COL[cat_name]
         cl = get_column_letter(col)
         ws.cell(row=2, column=col, value=f"=SUM({cl}7:{cl}{last_row})")
     if last_row >= 7:
         ws.cell(row=2, column=9, value=f"=SUM(I7:I{last_row})")

def find_sheet_by_name(wb, target_name):
     target_stripped = target_name.strip()
     for existing_name in wb.sheetnames:
         if existing_name.strip() == target_stripped:
             return existing_name, wb[existing_name]
     return None, None


# ============================================================
# 主流程 - process_monthly_file
# ============================================================

def process_monthly_file(dist_groups, dist_info, month_key, existing_bytes=None):
     year_str, mon_str = month_key.split("-")
     month_int = int(mon_str)
     month_display = f"{month_int}月份"
     output_filename = f"04_客戶對帳單{year_str}-{mon_str}.xlsx"

     m_year = int(year_str) - 1911
     month_title_r1 = f"{m_year}年{int(mon_str):02d}月"

     stats = {"added": 0, "skipped": 0, "dist_count": len(dist_groups)}
     log_lines = []
     def log(msg):
         log_lines.append(msg)

      # Build std_name -> sheet_name map from dist_info
     sheet_to_std = {}
     for sn, info in dist_info.items():
         sname_map = info.get("sheet_name", sn)
         sheet_to_std[sname_map] = sn

     if existing_bytes is not None:
         log("[既有檔案] 追加模式")
         try:
             existing_wb = openpyxl.load_workbook(io.BytesIO(existing_bytes))
         except Exception as e:
             log(f"[!] 讀取既有檔案失敗 ({e})，改為新建")
             existing_wb = None
     else:
         log("[新檔] 新建模式")
         existing_wb = None

     if existing_wb is None:
          # === 新建模式 ===
         wb = openpyxl.Workbook()
         ws_summary = wb.active
         ws_summary.title = "彙總"
         _init_summary_sheet(ws_summary, dist_info)
         ws_summary.cell(row=1, column=16, value=month_title_r1).number_format = "@"

         for sname in SHEET_ORDER:
             std_name_for_sname = SHEET_TO_STD.get(sname, sname)
             records = dist_groups.get(std_name_for_sname, {})
             ws = create_new_distributor_sheet(wb, std_name_for_sname, dist_info, month_display)

             if records:
                 new_last, added, skipped = append_data_to_sheet(ws, records)
                 update_sheet_formulas(ws, new_last)
                 stats["added"] += added
                 stats["skipped"] += skipped
                 log(f"[新建] {sname} - {added} 筆")
             else:
                 log(f"[新建] {sname} - 無資料（空頁）")

     else:
          # === 追加模式 ===
         wb = existing_wb
         if "彙總" not in wb.sheetnames:
             ws_summary_new = wb.create_sheet("彙總", 0)
             _init_summary_sheet(ws_summary_new, dist_info)
             ws_summary_new.cell(row=1, column=16, value=month_title_r1).number_format = "@"

         for std_name_in_groups in sorted(dist_groups.keys()):
             records = dist_groups[std_name_in_groups]
             info = dist_info.get(std_name_in_groups, {})
             sname_lookup = info.get("sheet_name", std_name_in_groups)

              # Match to SHEET_ORDER
             matched_sname = None
             for so in SHEET_ORDER:
                 if safe_sheet_name(so).strip() == safe_sheet_name(sname_lookup).strip():
                     matched_sname = so
                     break
             if not matched_sname:
                 matched_sname = safe_sheet_name(sname_lookup)

             found_name, ws = find_sheet_by_name(wb, matched_sname)

             if found_name and found_name != matched_sname:
                 log(f"[修正] {std_name_in_groups} - Sheet '{found_name}' -> '{matched_sname}'")
                 del wb[found_name]
                 ws = create_new_distributor_sheet(wb, std_name_in_groups, dist_info, month_display)
                 new_last, added, skipped = append_data_to_sheet(ws, records)
                 update_sheet_formulas(ws, new_last)
             elif found_name:
                 new_last, added, skipped = append_data_to_sheet(ws, records)
                 if added > 0:
                     update_sheet_formulas(ws, new_last)
             else:
                 log(f"[新頁] {std_name_in_groups} ({matched_sname})")
                 ws = create_new_distributor_sheet(wb, std_name_in_groups, dist_info, month_display)
                 new_last, added, skipped = append_data_to_sheet(ws, records)
                 update_sheet_formulas(ws, new_last)

             stats["added"] += added
             stats["skipped"] += skipped
             log(f"[追加] {std_name_in_groups} - {added} 筆新增, {skipped} 筆跳過")

     output_buf = io.BytesIO()
     wb.save(output_buf)
     output_bytes = output_buf.getvalue()
     wb.close()

     return output_bytes, output_filename, stats, log_lines
