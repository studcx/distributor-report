# -*- coding: utf-8 -*-
"""
init_db.py v4.2 - 資料庫管理模組
v4.2 變更：修正 get_available_months 的 conn.cursor() 重複呼叫問題、補上所有遺缺函式
"""
import sqlite3, os, sys
from datetime import date, datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

import openpyxl
from config import DB_PATH, CATEGORY_MAP, CATEGORY_ORDER, MANUAL_CODE_MAPPING, NAME_OVERRIDE, DEFAULT_DISCOUNT_RATES


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def get_category(product_code):
    if not product_code or len(str(product_code)) < 2:
        return "其他"
    prefix = str(product_code)[:2]
    return CATEGORY_MAP.get(prefix, "其他")


def calc_accounting_month(sale_date_str):
     if not sale_date_str or len(sale_date_str) < 7:
         return sale_date_str
     try:
         year = int(sale_date_str[:4])
         month = int(sale_date_str[5:7])
         day = int(sale_date_str[8:10]) if len(sale_date_str) >= 10 else 1
         if day >= 26:
             month += 1
             if month > 12:
                 month = 1
                 year += 1
         return f"{year:04d}-{month:02d}"
     except (ValueError, IndexError):
         return sale_date_str


def ensure_db(conn):
    cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS customer_mapping (
        id INTEGER PRIMARY KEY AUTOINCREMENT, cust_code TEXT NOT NULL UNIQUE,
        cust_short_name TEXT, cust_full_name TEXT, collector_code TEXT,
        collector_name TEXT, distributor_name TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS distributor_info (
        id INTEGER PRIMARY KEY AUTOINCREMENT, std_name TEXT NOT NULL UNIQUE,
        sheet_name TEXT, company_full TEXT, rate_dan_jin REAL DEFAULT 0.8,
        rate_fu_jin REAL DEFAULT 0.5, rate_dan_nong REAL DEFAULT 0.5,
        rate_fu_nong REAL DEFAULT 0.45, rate_otc REAL DEFAULT 0.5,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS import_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT, source_file TEXT, action TEXT,
        records_cust INTEGER DEFAULT 0, records_dist INTEGER DEFAULT 0,
        note TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS sales_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sale_date TEXT NOT NULL, invoice_no TEXT NOT NULL,
        cust_code TEXT, cust_name TEXT, product_code TEXT,
        category TEXT, amount REAL DEFAULT 0, distributor TEXT,
        input_by TEXT, accounting_month TEXT, batch_id INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS import_batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        input_by TEXT, total_lines INTEGER DEFAULT 0,
        new_records INTEGER DEFAULT 0, skipped_duplicates INTEGER DEFAULT 0,
        source TEXT, month_key TEXT)""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS user_profiles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE, full_name TEXT,
        is_active INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    conn.commit()


def _ensure_v4_tables(conn):
    cursor = conn.cursor()
    cur_check = conn.cursor()
    # Migrate: add accounting_month column if not exists
    cur_check.execute("PRAGMA table_info(sales_records)")
    col_names = [r[1] for r in cur_check.fetchall()]
    if "accounting_month" not in col_names:
        cursor.execute("ALTER TABLE sales_records ADD COLUMN accounting_month TEXT")
        # Backfill existing records with natural month as fallback
        rows = cursor.execute("SELECT id, sale_date FROM sales_records WHERE accounting_month IS NULL AND sale_date IS NOT NULL").fetchall()
        for rid, sd in rows:
            am = calc_accounting_month(sd)
            cursor.execute("UPDATE sales_records SET accounting_month = ? WHERE id = ?", (am, rid))
        conn.commit()
     # Migrate: add order_note column if not exists
    cur_check3 = conn.cursor()
    cur_check3.execute("PRAGMA table_info(sales_records)")
    col_names3 = [r[1] for r in cur_check3.fetchall()]
    if "order_note" not in col_names3:
        cursor.execute("ALTER TABLE sales_records ADD COLUMN order_note TEXT DEFAULT ''")
        conn.commit()
    cur_check2 = conn.cursor()
    cur_check.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='DIST_DISCOUNTS'")
    if not cur_check.fetchone():
        cur_check.execute("""CREATE TABLE DIST_DISCOUNTS (
                DISTINCT_NAME TEXT NOT NULL UNIQUE,
                DISCOUNT_VALUE REAL DEFAULT 0.8)""")
        rows = conn.execute("SELECT std_name, rate_dan_jin FROM distributor_info").fetchall()
        for r in rows:
            cur_check.execute("INSERT OR IGNORE INTO DIST_DISCOUNTS (DISTINCT_NAME, DISCOUNT_VALUE) VALUES (?, ?)", (r[0], r[1]))
    cur_check.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='CODE_TO_DIST'")
    if not cur_check.fetchone():
        cur_check.execute("""CREATE TABLE CODE_TO_DIST (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                CUSTOMER_CODE TEXT NOT NULL,
                CUSTOMER_NAME TEXT,
                DISTINCT_NAME TEXT NOT NULL,
                SHEET_NAME TEXT,
                IS_ACTIVE INTEGER DEFAULT 1)""")
        rows = conn.execute("SELECT cust_code, cust_short_name, cust_full_name, distributor_name FROM customer_mapping").fetchall()
        for r in rows:
            cur_check.execute("INSERT INTO CODE_TO_DIST (CUSTOMER_CODE, CUSTOMER_NAME, DISTINCT_NAME, SHEET_NAME) VALUES (?, ?, ?, ?)",
                               (r[0], r[2] or r[1], r[3], r[1]))

    # Migrate: add distributor_notes table if not exists
    cur_check.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='distributor_notes'")
    if not cur_check.fetchone():
        cur_check.execute("""CREATE TABLE distributor_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                distributor TEXT NOT NULL,
                month_key TEXT NOT NULL,
                note_text TEXT DEFAULT '',
                UNIQUE(distributor, month_key))""")

    conn.commit()

def ensure_user(conn, username):
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM user_profiles WHERE username = ?", (username,))
    if not cursor.fetchone():
        cursor.execute("INSERT INTO user_profiles (username, full_name) VALUES (?, ?)", (username, username))
    conn.commit()


def insert_sales_records(records_list, input_by, batch_info):
    if not records_list:
        return {"new": 0, "skipped": 0}
    conn = sqlite3.connect(DB_PATH)
    ensure_db(conn)
    cursor = conn.cursor()
    batch_id = batch_info.get("id", None)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_count = 0
    skipped_count = 0
    for rec in records_list:
        invoice_no = rec.get("invoice_no", "") or ""
        d_val = rec.get("date")
        if d_val and isinstance(d_val, (datetime, date)):
            sale_date = d_val.strftime("%Y-%m-%d")
        else:
            sale_date = str(d_val) if d_val else now_str[:10]
        cursor.execute("SELECT id FROM sales_records WHERE invoice_no = ? AND cust_code = ? AND category = ?",
                       (invoice_no, rec.get("cust_code", ""), rec.get("category", "")))
        if cursor.fetchone():
            skipped_count += 1
            continue
        acct_month = calc_accounting_month(sale_date)
        cursor.execute("INSERT INTO sales_records (sale_date, invoice_no, cust_code, cust_name, product_code, category, amount, distributor, input_by, accounting_month, batch_id, order_note) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                       (sale_date, invoice_no, rec.get("cust_code", ""), rec.get("cust_name", ""),
                        rec.get("product_code", ""), rec.get("category", "其他"),
                        rec.get("amount", 0.0), rec.get("distributor", ""), input_by, acct_month, batch_id,
                        rec.get("order_note", "") or ""))
        new_count += 1
    if batch_id:
        cursor.execute("UPDATE import_batches SET new_records = ?, skipped_duplicates = ? WHERE id = ?", (new_count, skipped_count, batch_id))
    conn.commit()
    conn.close()
    return {"new": new_count, "skipped": skipped_count}


def get_sales_by_month(conn, month_key):
    cursor = conn.cursor()
    cursor.execute("SELECT sale_date, invoice_no, cust_code, cust_name, product_code, category, amount, distributor, input_by, COALESCE(order_note,'') FROM sales_records WHERE accounting_month = ? ORDER BY sale_date, invoice_no", (month_key,))
    rows = cursor.fetchall()
    result = []
    for r in rows:
        result.append({"date": r[0], "invoice_no": r[1], "cust_code": r[2], "cust_name": r[3],
                       "product_code": r[4], "category": r[5], "amount": r[6], "distributor": r[7], "input_by": r[8], "order_note": r[9]})
    return result


def get_available_months(conn):
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT accounting_month as month_key FROM sales_records WHERE accounting_month IS NOT NULL AND accounting_month != '' GROUP BY month_key ORDER BY month_key DESC")
    months = [r[0] for r in cursor.fetchall()]
    return months


def delete_month_data(conn, month_key, input_by):
    cursor = conn.cursor()
    cursor.execute("DELETE FROM sales_records WHERE accounting_month = ?", (month_key,))
    deleted = cursor.rowcount
    conn.commit()
    return deleted


def get_recent_imports(conn, limit=50):
    cursor = conn.cursor()
    cursor.execute("SELECT id, batch_time, input_by, total_lines, new_records, skipped_duplicates, month_key FROM import_batches ORDER BY batch_time DESC LIMIT ?", (limit,))
    return cursor.fetchall()


def get_daily_trend(conn, month_key):
    cursor = conn.cursor()
    cursor.execute("SELECT substr(sale_date, 1, 10) as day, SUM(amount) as total FROM sales_records WHERE accounting_month = ? GROUP BY day ORDER BY day", (month_key,))
    return cursor.fetchall()


def get_distributor_ranking(conn, month_key):
    cursor = conn.cursor()
    cursor.execute("SELECT distributor, COUNT(*) as cnt, SUM(amount) as total FROM sales_records WHERE accounting_month = ? AND distributor IS NOT NULL AND distributor != '' GROUP BY distributor ORDER BY total DESC", (month_key,))
    return cursor.fetchall()


def get_category_breakdown(conn, month_key):
    cursor = conn.cursor()
    cursor.execute("SELECT category, SUM(amount) as total FROM sales_records WHERE accounting_month = ? GROUP BY category ORDER BY total DESC", (month_key,))
    return cursor.fetchall()


def get_overview_stats(conn, month_key):
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*), COALESCE(SUM(amount),0), COUNT(DISTINCT invoice_no), COUNT(DISTINCT distributor) FROM sales_records WHERE accounting_month = ?", (month_key,))
    row = cursor.fetchone()
    return {"total_lines": row[0], "total_amount": row[1], "invoice_count": row[2], "distributor_count": row[3]}




def import_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    conn = sqlite3.connect(DB_PATH)
    ensure_db(conn)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM customer_mapping")
    cursor.execute("DELETE FROM distributor_info")
    ws_cust = wb["客戶對照表"]
    cust_count = 0
    for row in ws_cust.iter_rows(min_row=2, max_col=5):
        code = str(row[0].value).strip() if row[0].value else ""
        if not code: continue
        short_name = str(row[1].value).strip() if row[1].value else ""
        full_name = str(row[2].value).strip() if row[2].value else ""
        collector_code = str(row[3].value).strip() if row[3].value else ""
        collector_name_raw = str(row[4].value).strip() if row[4].value else ""
        if not collector_name_raw or collector_name_raw == "#N/A":
            dist_name = MANUAL_CODE_MAPPING.get(code)
            if not dist_name: continue
        else:
            dist_name = NAME_OVERRIDE.get(collector_name_raw, collector_name_raw)
        cursor.execute("INSERT OR REPLACE INTO customer_mapping (cust_code, cust_short_name, cust_full_name, collector_code, collector_name, distributor_name) VALUES (?, ?, ?, ?, ?, ?)",
              (code, short_name, full_name, collector_code, collector_name_raw, dist_name))
        cust_count += 1
    ws_disc = wb["經銷商折數基準表"]
    dist_count = 0
    for row in ws_disc.iter_rows(min_row=2, max_col=8):
        std_name_raw = row[0].value
        if not std_name_raw: continue
        std_name = str(std_name_raw).strip()
        sheet_name = str(row[1].value) if row[1].value is not None else std_name
        company_full = str(row[2].value).strip() if row[2].value else std_name
        cursor.execute("INSERT OR REPLACE INTO distributor_info (std_name, sheet_name, company_full, rate_dan_jin, rate_fu_jin, rate_dan_nong, rate_fu_nong, rate_otc) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
              (std_name, sheet_name, company_full, safe_float(row[3].value), safe_float(row[4].value),
               safe_float(row[5].value), safe_float(row[6].value), safe_float(row[7].value)))
        dist_count += 1
    cursor.execute("INSERT INTO import_log (source_file, action, records_cust, records_dist, note) VALUES (?, 'import', ?, ?, ?)",
          (file_path, cust_count, dist_count, f"Import from {os.path.basename(file_path)}"))
    conn.commit()

def _get_discount_rates_dict(conn):
    """Return dict of {distributor_name: {category: rate}} for actual price calculation."""
    cursor = conn.cursor()
    cursor.execute("SELECT std_name, rate_dan_jin, rate_fu_jin, rate_dan_nong, rate_fu_nong, rate_otc FROM distributor_info")
    rates_map = {}
    for row in cursor.fetchall():
        dn = row[0]
        rates_map[dn] = {
            "單斤": row[1] if row[1] is not None else 0.8,
            "複斤": row[2] if row[2] is not None else 0.5,
            "單濃": row[3] if row[3] is not None else 0.5,
            "複濃": row[4] if row[4] is not None else 0.45,
            "O.T.C.": row[5] if row[5] is not None else 0.5,
        }
    return rates_map


def _apply_discount(amount, category, distributor, rates_map):
    """Apply discount rate for a single record based on its category and distributor."""
    rmx = rates_map.get(distributor, {})
    rate = rmx.get(category, 1.0)
    if category == "其他":
        return amount  # no discount for 其他
    return round(amount * rate)


def get_dashboard_data(conn, month_key):
    """Unified dashboard data: returns brand AND actual prices for all charts."""
    cursor = conn.cursor()
    cursor.execute("SELECT sale_date, invoice_no, distributor, category, SUM(amount) as total FROM sales_records WHERE accounting_month = ? GROUP BY sale_date, invoice_no, distributor, category ORDER BY sale_date", (month_key,))
    raw_rows = cursor.fetchall()

    # Load discount rates
    rates_map = _get_discount_rates_dict(conn)

    # === 每日銷售趨勢：每個日期同時計算牌價與實價 ===
    daily_agg = {}
    for row in raw_rows:
        day = str(row[0])[:10] if row[0] else ""
        cat = row[3] or "其他"
        amt = row[4] or 0.0
        dist = row[2] or ""
        if not day:
            continue
        if day not in daily_agg:
            daily_agg[day] = {"brand": 0.0, "actual": 0.0}
        daily_agg[day]["brand"] += amt
        daily_agg[day]["actual"] += _apply_discount(amt, cat, dist, rates_map)

    daily_data = []
    for day in sorted(daily_agg.keys()):
        daily_data.append({
            "date": day,
            "brand": round(daily_agg[day]["brand"]),
            "actual": round(daily_agg[day]["actual"]),
        })

    # === 經銷商排名：同時計算牌價與實價，使用銷貨單張數 ===
    dist_agg = {}
    for row in raw_rows:
        dn = row[2] or ""
        cat = row[3] or "其他"
        amt = row[4] or 0.0
        inv_no = row[1] or ""
        if not dn:
            continue
        if dn not in dist_agg:
            dist_agg[dn] = {"brand": 0.0, "actual": 0.0, "invoices": set()}
        dist_agg[dn]["brand"] += amt
        dist_agg[dn]["actual"] += _apply_discount(amt, cat, dn, rates_map)
        dist_agg[dn]["invoices"].add(inv_no)

    dist_data = []
    for dn, info in sorted(dist_agg.items(), key=lambda x: x[1]["brand"], reverse=True):
        dist_data.append({
            "distributor": dn,
            "brand": round(info["brand"]),
            "actual": round(info["actual"]),
            "invoice_count": len(info["invoices"]),
        })

    # === 產品類別佔比：同時計算牌價與實價 ===
    cat_agg = {}
    for row in raw_rows:
        cat = row[3] or "其他"
        amt = row[4] or 0.0
        dist = row[2] or ""
        if cat not in cat_agg:
            cat_agg[cat] = {"brand": 0.0, "actual": 0.0}
        cat_agg[cat]["brand"] += amt
        cat_agg[cat]["actual"] += _apply_discount(amt, cat, dist, rates_map)

    cat_data = []
    for cat in ["單斤", "複斤", "單濃", "複濃", "O.T.C.", "其他"]:
        if cat in cat_agg:
            cat_data.append({
                "category": cat,
                "brand": round(cat_agg[cat]["brand"]),
                "actual": round(cat_agg[cat]["actual"]),
            })

    # === 總覽統計 ===
    total_brand = sum(r["brand"] for r in dist_data) if dist_data else 0
    total_actual = sum(r["actual"] for r in dist_data) if dist_data else 0
    cursor.execute("SELECT COUNT(*), COUNT(DISTINCT invoice_no) FROM sales_records WHERE accounting_month = ?", (month_key,))
    row_count = cursor.fetchone()

    stats = {
        "total_lines": row_count[0] if row_count else 0,
        "total_brand": round(total_brand),
        "total_actual": round(total_actual),
        "invoice_count": row_count[1] if row_count else 0,
        "distributor_count": len(dist_data),
    }

    return {
        "daily": daily_data,
        "distributors": dist_data,
        "categories": cat_data,
        "stats": stats,
    }


def import_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    conn = sqlite3.connect(DB_PATH)
    ensure_db(conn)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM customer_mapping")
    cursor.execute("DELETE FROM distributor_info")
    ws_cust = wb["客戶對照表"]
    cust_count = 0
    for row in ws_cust.iter_rows(min_row=2, max_col=5):
        code = str(row[0].value).strip() if row[0].value else ""
        if not code: continue
        short_name = str(row[1].value).strip() if row[1].value else ""
        full_name = str(row[2].value).strip() if row[2].value else ""
        collector_code = str(row[3].value).strip() if row[3].value else ""
        collector_name_raw = str(row[4].value).strip() if row[4].value else ""
        if not collector_name_raw or collector_name_raw == "#N/A":
            dist_name = MANUAL_CODE_MAPPING.get(code)
            if not dist_name: continue
        else:
            dist_name = NAME_OVERRIDE.get(collector_name_raw, collector_name_raw)
        cursor.execute("INSERT OR REPLACE INTO customer_mapping (cust_code, cust_short_name, cust_full_name, collector_code, collector_name, distributor_name) VALUES (?, ?, ?, ?, ?, ?)",
              (code, short_name, full_name, collector_code, collector_name_raw, dist_name))
        cust_count += 1
    ws_disc = wb["經銷商折數基準表"]
    dist_count = 0
    for row in ws_disc.iter_rows(min_row=2, max_col=8):
        std_name_raw = row[0].value
        if not std_name_raw: continue
        std_name = str(std_name_raw).strip()
        sheet_name = str(row[1].value) if row[1].value is not None else std_name
        company_full = str(row[2].value).strip() if row[2].value else std_name
        cursor.execute("INSERT OR REPLACE INTO distributor_info (std_name, sheet_name, company_full, rate_dan_jin, rate_fu_jin, rate_dan_nong, rate_fu_nong, rate_otc) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
              (std_name, sheet_name, company_full, safe_float(row[3].value), safe_float(row[4].value),
               safe_float(row[5].value), safe_float(row[6].value), safe_float(row[7].value)))
        dist_count += 1
    cursor.execute("INSERT INTO import_log (source_file, action, records_cust, records_dist, note) VALUES (?, 'import', ?, ?, ?)",
          (file_path, cust_count, dist_count, f"Import from {os.path.basename(file_path)}"))
    conn.commit()

def _get_discount_rates_dict(conn):
    """Return dict of {distributor_name: {category: rate}} for actual price calculation."""
    cursor = conn.cursor()
    cursor.execute("SELECT std_name, rate_dan_jin, rate_fu_jin, rate_dan_nong, rate_fu_nong, rate_otc FROM distributor_info")
    rates_map = {}
    for row in cursor.fetchall():
        dn = row[0]
        rates_map[dn] = {
            "單斤": row[1] if row[1] is not None else 0.8,
            "複斤": row[2] if row[2] is not None else 0.5,
            "單濃": row[3] if row[3] is not None else 0.5,
            "複濃": row[4] if row[4] is not None else 0.45,
            "O.T.C.": row[5] if row[5] is not None else 0.5,
        }
    return rates_map


def _apply_discount(amount, category, distributor, rates_map):
    """Apply discount rate for a single record based on its category and distributor."""
    rmx = rates_map.get(distributor, {})
    rate = rmx.get(category, 1.0)
    if category == "其他":
        return amount  # no discount for 其他
    return round(amount * rate)


def get_daily_trend_with_price_mode(conn, month_key, price_mode):
    """Daily trend supporting both brand price and actual price."""
    cursor = conn.cursor()
    cursor.execute("SELECT substr(sale_date,1,10) as day, distributor, category, SUM(amount) as total FROM sales_records WHERE accounting_month = ? GROUP BY day, distributor, category ORDER BY day", (month_key,))
    daily_agg = {}
    for row in cursor.fetchall():
        day, dist, cat, amt = row[0], row[1] or "", row[4] if len(row) > 4 else "其他", row[3] or 0.0
        key = day
        if key not in daily_agg:
            daily_agg[key] = []
        daily_agg[key].append({"dist": dist, "cat": cat, "amt": amt})

    if price_mode == "實價":
        rates_map = _get_discount_rates_dict(conn)
    result = []
    for day in sorted(daily_agg.keys()):
        total = 0
        for item in daily_agg[day]:
            if price_mode == "實價":
                total += _apply_discount(item["amt"], item["cat"], item["dist"], rates_map)
            else:
                total += item["amt"]
        result.append((day, total))
    return result


def get_distributor_ranking_with_price_mode(conn, month_key, price_mode):
    """Distributor ranking supporting both brand price and actual price."""
    cursor = conn.cursor()
    cursor.execute("SELECT distributor, category, SUM(amount) as total FROM sales_records WHERE accounting_month = ? AND distributor IS NOT NULL AND distributor != '' GROUP BY distributor, category ORDER BY total DESC", (month_key,))

    if price_mode == "實價":
        rates_map = _get_discount_rates_dict(conn)

    dist_agg = {}
    for row in cursor.fetchall():
        dn, cat, amt = row[0], row[2] if len(row) > 2 else "其他", row[2]
        # Re-parse: distributor=col0, category=col1, total=col2
        dn, cat, amt = row[0], row[1] or "其他", row[2] or 0.0
        if dn not in dist_agg:
            dist_agg[dn] = {"cnt": 0, "amt": 0}
        dist_agg[dn]["cnt"] += 1
        if price_mode == "實價":
            dist_agg[dn]["amt"] += _apply_discount(amt, cat, dn, rates_map)
        else:
            dist_agg[dn]["amt"] += amt

    result = []
    for dn, info in dist_agg.items():
        result.append((dn, info["cnt"], round(info["amt"])))
    return result


def get_category_breakdown_with_price_mode(conn, month_key, price_mode):
    """Category breakdown supporting both brand price and actual price."""
    cursor = conn.cursor()
    cursor.execute("SELECT category, distributor, SUM(amount) as total FROM sales_records WHERE accounting_month = ? GROUP BY category, distributor ORDER BY total DESC", (month_key,))

    if price_mode == "實價":
        rates_map = _get_discount_rates_dict(conn)

    cat_agg = {}
    for row in cursor.fetchall():
        cat, dn, amt = row[0] or "其他", row[1] or "", row[2] or 0.0
        if cat not in cat_agg:
            cat_agg[cat] = 0
        if price_mode == "實價":
            cat_agg[cat] += _apply_discount(amt, cat, dn, rates_map)
        else:
            cat_agg[cat] += amt

    result = []
    for cat, total in cat_agg.items():
        result.append((cat, round(total)))
    return result


def get_overview_stats_with_price_mode(conn, month_key, price_mode):
    """Overview stats supporting both brand price and actual price."""
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*), COUNT(DISTINCT invoice_no), COUNT(DISTINCT distributor) FROM sales_records WHERE accounting_month = ?", (month_key,))
    row = cursor.fetchone()

    if price_mode == "實價":
        rates_map = _get_discount_rates_dict(conn)
        cursor2 = conn.cursor()
        cursor2.execute("SELECT distributor, category, SUM(amount) as total FROM sales_records WHERE accounting_month = ? GROUP BY distributor, category", (month_key,))
        total_amt = 0
        for r in cursor2.fetchall():
            dn, cat, amt = r[0] or "", r[1] or "其他", r[2] or 0.0
            total_amt += _apply_discount(amt, cat, dn, rates_map)
    else:
        cursor2 = conn.cursor()
        cursor2.execute("SELECT COALESCE(SUM(amount),0) FROM sales_records WHERE accounting_month = ?", (month_key,))
        total_amt = cursor2.fetchone()[0] or 0.0

    return {"total_lines": row[0], "total_amount": round(total_amt), "invoice_count": row[1], "distributor_count": row[2]}

    wb = openpyxl.load_workbook(file_path)
    conn = sqlite3.connect(DB_PATH)
    ensure_db(conn)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM customer_mapping")
    cursor.execute("DELETE FROM distributor_info")
    ws_cust = wb["客戶對照表"]
    cust_count = 0
    for row in ws_cust.iter_rows(min_row=2, max_col=5):
        code = str(row[0].value).strip() if row[0].value else ""
        if not code: continue
        short_name = str(row[1].value).strip() if row[1].value else ""
        full_name = str(row[2].value).strip() if row[2].value else ""
        collector_code = str(row[3].value).strip() if row[3].value else ""
        collector_name_raw = str(row[4].value).strip() if row[4].value else ""
        if not collector_name_raw or collector_name_raw == "#N/A":
            dist_name = MANUAL_CODE_MAPPING.get(code)
            if not dist_name: continue
        else:
            dist_name = NAME_OVERRIDE.get(collector_name_raw, collector_name_raw)
        cursor.execute("INSERT OR REPLACE INTO customer_mapping (cust_code, cust_short_name, cust_full_name, collector_code, collector_name, distributor_name) VALUES (?, ?, ?, ?, ?, ?)",
              (code, short_name, full_name, collector_code, collector_name_raw, dist_name))
        cust_count += 1
    ws_disc = wb["經銷商折數基準表"]
    dist_count = 0
    for row in ws_disc.iter_rows(min_row=2, max_col=8):
        std_name_raw = row[0].value
        if not std_name_raw: continue
        std_name = str(std_name_raw).strip()
        sheet_name = str(row[1].value) if row[1].value is not None else std_name
        company_full = str(row[2].value).strip() if row[2].value else std_name
        cursor.execute("INSERT OR REPLACE INTO distributor_info (std_name, sheet_name, company_full, rate_dan_jin, rate_fu_jin, rate_dan_nong, rate_fu_nong, rate_otc) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
              (std_name, sheet_name, company_full, safe_float(row[3].value), safe_float(row[4].value),
               safe_float(row[5].value), safe_float(row[6].value), safe_float(row[7].value)))
        dist_count += 1
    cursor.execute("INSERT INTO import_log (source_file, action, records_cust, records_dist, note) VALUES (?, 'import', ?, ?, ?)",
          (file_path, cust_count, dist_count, f"Import from {os.path.basename(file_path)}"))
    conn.commit()
    wb.close()
    ensure_user(conn, "system")
    return cust_count, dist_count


def load_mappings_from_db():
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError("Database not found.")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT cust_code, distributor_name FROM customer_mapping")
    code_to_dist = {row[0]: row[1] for row in cursor.fetchall()}
    cursor.execute("SELECT std_name, sheet_name, company_full, rate_dan_jin, rate_fu_jin, rate_dan_nong, rate_fu_nong, rate_otc FROM distributor_info")
    dist_info = {}
    for row in cursor.fetchall():
        sn = row[1] if row[1] is not None else row[0]
        cf = row[2] if row[2] is not None else row[0]
        dist_info[row[0]] = {"sheet_name": sn, "company_full": cf,
            "discount_rates": {"單斤": row[3] if row[3] is not None else 0.8, "複斤": row[4] if row[4] is not None else 0.5,
                                "單濃": row[5] if row[5] is not None else 0.5, "複濃": row[6] if row[6] is not None else 0.45,
                                "O.T.C.": row[7] if row[7] is not None else 0.5}}
    conn.close()
    return code_to_dist, dist_info


def get_distributor_list():
    if not os.path.exists(DB_PATH):
        return []
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT std_name FROM distributor_info ORDER BY std_name")
    result = [row[0] for row in cursor.fetchall()]
    conn.close()
    return result


if __name__ == "__main__":
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = os.path.expanduser("~/Desktop/Temp/02_經銷商對應與篩選對照表.xlsm")
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        sys.exit(1)
    try:
        import_from_excel(file_path)
    except Exception as e:
        print(f"Import failed: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)


def get_distributor_note(conn, distributor, month_key):
    cursor = conn.cursor()
    cursor.execute("SELECT note_text FROM distributor_notes WHERE distributor = ? AND month_key = ?", (distributor, month_key))
    row = cursor.fetchone()
    return row[0] if row and row[0] else ''


def save_distributor_note(conn, distributor, month_key, note_text):
    cursor = conn.cursor()
    cursor.execute("INSERT INTO distributor_notes (distributor, month_key, note_text) VALUES (?, ?, ?) ON CONFLICT(distributor, month_key) DO UPDATE SET note_text = ?", (distributor, month_key, note_text, note_text))
    conn.commit()

